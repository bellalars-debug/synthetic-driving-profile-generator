"""Human-readable .xlsx export layer used for research validation.

Cross-cutting: not a scientific pipeline stage itself. Reads the finished
pipeline artifacts (`data/processed/synthetic_employees.parquet`,
`data/processed/synthetic_activity.parquet`) plus the project's own
validation docs (`docs/validation_results.md`, `docs/model_status.md`) and
writes one reviewable workbook (`reports/xlsx/synthetic_mobility_report.xlsx`)
so a researcher or company stakeholder can inspect the synthetic population
without Python or Parquet tooling. No generation logic, scientific
assumption, or validation threshold lives here or is changed by importing
this module - it is a read-only rendering layer. EV ownership/charging
demand is out of scope (see `scenarios/charging_demand.py`) and is not
referenced here beyond a note that it is not yet included.

## Workbook contents

1. **README** - purpose, generation date, row counts, source dataset, a
   plain-language "this is synthetic data" / "no EV assumptions yet" note,
   and a one-line description of every other sheet.
2. **Synthetic Employees** - one row per synthetic employee, restricted to
   the caller-facing columns listed in `EMPLOYEE_EXPORT_COLUMNS` (excludes
   `source_houseid`/`source_personid` - real NHTS traceability IDs that
   should not ship in a public-facing workbook, per `generator/sample.py`'s
   own module docstring on why those columns exist at all).
3. **Synthetic Activity** - one row per trip leg, restricted to
   `ACTIVITY_EXPORT_COLUMNS` (excludes `vehicle_type`/`vehicle_fuel`, which
   the task's column list does not request).
4. **Employee Summary** - population-level descriptive statistics plus a
   cluster count/share breakdown.
5. **Cluster Summary** - the same descriptive statistics broken out by
   `cluster_id`.
6. **Hourly Workplace Presence** - estimated employee count/share present
   at the workplace during each hour of the day, derived from each
   workplace-arrival leg's `arrival_time` + `workplace_dwell_minutes`
   window (`compute_hourly_workplace_presence`).
7. **Validation Summary** - a small set of validation metrics recomputed
   directly from the final output tables (fallback-chain rate, implausible-
   speed leg count - the same definitions `validation/activity.py` uses,
   reused via `driving_profiles.generator.activity`'s own plausibility
   constants rather than re-deriving them), plus the top-line pass/fail/
   informational counts and the documented remaining limitations pulled
   from `docs/validation_results.md`/`docs/model_status.md`. Deliberately
   does not re-parse every row of `validation_results.md`'s per-metric
   tables (`_extract_validation_totals` reads one summary line; parsing the
   full table would tightly and brittley couple this module to that
   report's exact column layout for no benefit a researcher needs here).
"""

from __future__ import annotations

import datetime as dt
import logging
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from driving_profiles.generator import activity as activity_module
from driving_profiles.generator.time_utils import MINUTES_PER_DAY, hhmm_to_minutes

logger = logging.getLogger(__name__)

DEFAULT_PROCESSED_DIR = Path("data/processed")
DEFAULT_REPORTS_DIR = Path("reports/xlsx")
DEFAULT_MODEL_STATUS_PATH = Path("docs/model_status.md")
DEFAULT_VALIDATION_RESULTS_PATH = Path("docs/validation_results.md")

REPORT_FILENAME = "synthetic_mobility_report.xlsx"

SOURCE_DATASET_LABEL = "NHTS 2022 (National Household Travel Survey)"

# --- Column selections (task spec: exactly these columns, in this order) ----

EMPLOYEE_EXPORT_COLUMNS = [
    "synthetic_employee_id",
    "cluster_id",
    "age",
    "age_band",
    "household_income_bracket",
    "household_size",
    "household_vehicle_count",
    "vehicles_per_driver",
    "commute_distance_survey_miles",
    "commute_distance_trip_miles",
    "commute_duration_minutes",
    "work_arrival_time",
    "work_departure_time",
    "trips_per_day",
    "total_daily_miles",
    "total_driving_minutes",
    "number_of_stops",
    "used_household_vehicle",
]

ACTIVITY_EXPORT_COLUMNS = [
    "synthetic_employee_id",
    "trip_number",
    "departure_time",
    "arrival_time",
    "trip_purpose",
    "distance",
    "duration",
    "dwell_time_after",
    "is_workplace_arrival",
    "is_workplace_departure",
    "workplace_dwell_minutes",
    "chain_source",
]

# HHMM-encoded columns (see generator/time_utils.py) rendered as Excel time
# values rather than raw numbers, so "8:30" reads as a time, not "830".
EMPLOYEE_TIME_COLUMNS = ["work_arrival_time", "work_departure_time"]
ACTIVITY_TIME_COLUMNS = ["departure_time", "arrival_time"]

# Explicit number formats for columns where type-based inference
# (_infer_number_format) would pick the wrong precision or shouldn't apply
# (a percent share, or an ID-like coded value best left as a plain integer).
EMPLOYEE_COLUMN_FORMATS = {
    "household_income_bracket": "0",
    "cluster_id": "0",
}

SHEET_DESCRIPTIONS = [
    ("README", "This sheet - report overview and a description of every other sheet."),
    (
        "Synthetic Employees",
        "One row per synthetic employee: demographics, household, commute, and "
        "daily driving summary features.",
    ),
    (
        "Synthetic Activity",
        "One row per trip leg in every synthetic employee's daily trip chain.",
    ),
    (
        "Employee Summary",
        "Population-level summary statistics across all synthetic employees.",
    ),
    (
        "Cluster Summary",
        "The same summary statistics broken out by behavioral cluster archetype.",
    ),
    (
        "Hourly Workplace Presence",
        "Estimated count and share of synthetic employees present at the "
        "workplace during each hour of the day.",
    ),
    (
        "Validation Summary",
        "Core validation metrics recomputed from the final output tables, plus "
        "documented remaining limitations.",
    ),
]

# --- Table styling ------------------------------------------------------------
# One distinct built-in style per sheet ("separate table-style formatting for
# each sheet" per the task spec) rather than one style reused everywhere.
TABLE_STYLES = [
    "TableStyleMedium2",
    "TableStyleMedium9",
    "TableStyleMedium15",
    "TableStyleMedium6",
    "TableStyleMedium11",
    "TableStyleMedium16",
    "TableStyleMedium4",
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

MAX_COLUMN_WIDTH = 42
MIN_COLUMN_WIDTH = 8


# --- Time-of-day helpers -----------------------------------------------------


def _minutes_to_time(minutes: float) -> dt.time | None:
    """Minutes-since-midnight -> `datetime.time`, for cells that should
    render with Excel's time formatting. `None` (an empty cell) for NaN,
    matching this project's "not applicable," not "midnight," convention
    for a missing time-of-day value.
    """
    if pd.isna(minutes):
        return None
    clamped = min(max(round(float(minutes)), 0), MINUTES_PER_DAY - 1)
    hours, mins = divmod(clamped, 60)
    return dt.time(hour=int(hours), minute=int(mins))


def _hhmm_to_time(hhmm: float) -> dt.time | None:
    """NHTS-style HHMM value -> `datetime.time` (see `time_utils.hhmm_to_minutes`
    for why this isn't a direct divmod on the raw HHMM number).
    """
    return _minutes_to_time(hhmm_to_minutes(hhmm))


# --- Loading -------------------------------------------------------------------


def load_synthetic_employees(processed_dir: Path = DEFAULT_PROCESSED_DIR) -> pd.DataFrame:
    """Read `generator/sample.py`'s output (`synthetic_employees.parquet`)."""
    return activity_module.load_synthetic_employees(processed_dir)


def load_synthetic_activity(processed_dir: Path = DEFAULT_PROCESSED_DIR) -> pd.DataFrame:
    """Read `generator/activity.py`'s output (`synthetic_activity.parquet`)."""
    path = Path(processed_dir) / activity_module.ACTIVITY_TABLE_FILENAME
    if not path.exists():
        raise FileNotFoundError(
            f"Synthetic activity table not found: {path}. Run "
            "`python -m driving_profiles.generator.activity` first."
        )
    return pd.read_parquet(path)


# --- Summary computation --------------------------------------------------------


def compute_employee_summary(
    employees: pd.DataFrame, activity: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Population-level descriptive statistics (`metric`/`value` rows) plus a
    separate cluster count/share table.

    Split into two frames rather than one, since a cluster breakdown has a
    different shape (one row per cluster, not one row per metric) than the
    single-population metrics - `write_workbook` renders each as its own
    Excel Table on the "Employee Summary" sheet.
    """
    n_employees = len(employees)
    arrival_minutes = employees["work_arrival_time"].apply(hhmm_to_minutes)
    departure_minutes = employees["work_departure_time"].apply(hhmm_to_minutes)
    workplace_dwell = activity.loc[activity["is_workplace_arrival"], "workplace_dwell_minutes"]

    metrics = pd.DataFrame(
        [
            {"metric": "Employee count", "value": n_employees},
            {
                "metric": "Average commute distance (mi)",
                "value": employees["commute_distance_survey_miles"].mean(),
            },
            {
                "metric": "Median commute distance (mi)",
                "value": employees["commute_distance_survey_miles"].median(),
            },
            {
                "metric": "Average commute duration (min)",
                "value": employees["commute_duration_minutes"].mean(),
            },
            {
                "metric": "Average work arrival time",
                "value": _minutes_to_time(arrival_minutes.mean()),
            },
            {
                "metric": "Average work departure time",
                "value": _minutes_to_time(departure_minutes.mean()),
            },
            {"metric": "Average trips per day", "value": employees["trips_per_day"].mean()},
            {"metric": "Average daily miles", "value": employees["total_daily_miles"].mean()},
            {
                "metric": "Average workplace dwell duration (min)",
                "value": workplace_dwell.mean(),
            },
        ]
    )

    cluster_counts = employees["cluster_id"].value_counts().sort_index()
    cluster_summary = pd.DataFrame(
        {
            "cluster_id": cluster_counts.index.astype("Int64"),
            "employee_count": cluster_counts.to_numpy(),
            "employee_share": (cluster_counts / n_employees).to_numpy() if n_employees else 0.0,
        }
    ).reset_index(drop=True)

    return metrics, cluster_summary


def compute_cluster_summary(employees: pd.DataFrame, activity: pd.DataFrame) -> pd.DataFrame:
    """One row per cluster: employee count/share and the driving-behavior
    means/medians requested by the task spec.

    Workplace dwell duration is computed from `activity` (not `employees`,
    which has no per-leg data) by joining each workplace-arrival leg back
    to its employee's `cluster_id`.
    """
    n_employees = len(employees)
    grouped = employees.groupby("cluster_id", dropna=True)
    summary = grouped.agg(
        employee_count=("synthetic_employee_id", "size"),
        mean_commute_distance_miles=("commute_distance_survey_miles", "mean"),
        median_commute_distance_miles=("commute_distance_survey_miles", "median"),
        mean_commute_duration_minutes=("commute_duration_minutes", "mean"),
        mean_trips_per_day=("trips_per_day", "mean"),
        mean_number_of_stops=("number_of_stops", "mean"),
        mean_total_daily_miles=("total_daily_miles", "mean"),
    ).reset_index()
    summary["employee_share"] = summary["employee_count"] / n_employees if n_employees else 0.0

    time_by_cluster = pd.DataFrame(
        {
            "cluster_id": employees["cluster_id"],
            "_arrival_min": employees["work_arrival_time"].apply(hhmm_to_minutes),
            "_departure_min": employees["work_departure_time"].apply(hhmm_to_minutes),
        }
    ).groupby("cluster_id", dropna=True).mean().reset_index()
    summary = summary.merge(time_by_cluster, on="cluster_id", how="left")
    summary["mean_arrival_time"] = summary["_arrival_min"].apply(_minutes_to_time)
    summary["mean_departure_time"] = summary["_departure_min"].apply(_minutes_to_time)
    summary = summary.drop(columns=["_arrival_min", "_departure_min"])

    employee_cluster_map = employees[["synthetic_employee_id", "cluster_id"]]
    workplace_legs = activity.loc[activity["is_workplace_arrival"]].merge(
        employee_cluster_map, on="synthetic_employee_id", how="left"
    )
    dwell_by_cluster = (
        workplace_legs.groupby("cluster_id", dropna=True)["workplace_dwell_minutes"]
        .mean()
        .rename("mean_workplace_dwell_minutes")
        .reset_index()
    )
    summary = summary.merge(dwell_by_cluster, on="cluster_id", how="left")

    column_order = [
        "cluster_id",
        "employee_count",
        "employee_share",
        "mean_commute_distance_miles",
        "median_commute_distance_miles",
        "mean_commute_duration_minutes",
        "mean_arrival_time",
        "mean_departure_time",
        "mean_trips_per_day",
        "mean_number_of_stops",
        "mean_total_daily_miles",
        "mean_workplace_dwell_minutes",
    ]
    return summary[column_order].sort_values("cluster_id").reset_index(drop=True)


def compute_hourly_workplace_presence(
    activity: pd.DataFrame, n_employees: int
) -> pd.DataFrame:
    """Estimated employee count/share present at the workplace during each
    hour of the day (0-23), from each workplace-arrival leg's
    `arrival_time` + `workplace_dwell_minutes` window.

    A leg with no recorded `workplace_dwell_minutes` (the day's diary ends
    before a departure leg is recorded - real in this project's population,
    see `generator/activity.py`'s module docstring) is treated as "present
    through the end of the day" rather than dropped, since the employee's
    own record has no evidence they left. An employee with more than one
    workplace-arrival leg that day (the fragmented-dwell-window case) is
    counted at most once per hour, not once per leg.
    """
    workplace_legs = activity.loc[activity["is_workplace_arrival"]].copy()
    start_min = workplace_legs["arrival_time"].apply(hhmm_to_minutes)
    dwell = workplace_legs["workplace_dwell_minutes"]
    end_min = (start_min + dwell.fillna(MINUTES_PER_DAY - start_min)).clip(upper=MINUTES_PER_DAY)
    workplace_legs = workplace_legs.assign(_start=start_min, _end=end_min)

    rows = []
    for hour in range(24):
        hour_start, hour_end = hour * 60, hour * 60 + 60
        overlap = (workplace_legs["_start"] < hour_end) & (workplace_legs["_end"] > hour_start)
        present = int(workplace_legs.loc[overlap, "synthetic_employee_id"].nunique())
        rows.append(
            {
                "hour": hour,
                "employees_present": present,
                "percent_present": present / n_employees if n_employees else 0.0,
            }
        )
    return pd.DataFrame(rows)


def compute_validation_metrics(activity: pd.DataFrame) -> pd.DataFrame:
    """Core structural validation metrics recomputed directly from the final
    `synthetic_activity` table, using `generator/activity.py`'s own
    plausibility constants (not re-derived here) - the same definitions
    `validation/activity.py` uses for these two checks.
    """
    per_employee = activity.drop_duplicates("synthetic_employee_id")
    n_employees = len(per_employee)
    n_fallback = int((per_employee["chain_source"] == activity_module.FALLBACK_CHAIN_SOURCE).sum())
    fallback_rate = n_fallback / n_employees if n_employees else 0.0

    legs = activity.loc[activity["duration"] > 0]
    implied_speed_mph = legs["distance"] / (legs["duration"] / 60.0)
    n_implausible = int(
        (
            (implied_speed_mph < activity_module.MIN_PLAUSIBLE_SPEED_MPH)
            | (implied_speed_mph > activity_module.MAX_PLAUSIBLE_SPEED_MPH)
        ).sum()
    )

    return pd.DataFrame(
        [
            {"metric": "Fallback chain rate", "value": fallback_rate},
            {"metric": "Implausible-speed leg count", "value": n_implausible},
            {"metric": "Legs checked for speed plausibility", "value": len(legs)},
        ]
    )


def _extract_validation_totals(text: str) -> tuple[int, int, int] | None:
    """Pull the single "**N passed, N failed, N informational**" summary
    line out of `docs/validation_results.md` - deliberately not a parse of
    that document's full per-metric tables (see module docstring).
    """
    match = re.search(r"\*\*(\d+) passed, (\d+) failed, (\d+) informational\*\*", text)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2)), int(match.group(3))


def _extract_markdown_bullets(text: str, heading_substring: str) -> list[str]:
    """Bullet lines (`- ...`) under the first `## ...` heading containing
    `heading_substring`, stopping at the next `## ` heading. Used to pull
    the "documented limitations" section out of `docs/model_status.md`
    without parsing the rest of that document.
    """
    bullets: list[str] = []
    in_section = False
    for line in text.splitlines():
        if line.startswith("## "):
            in_section = heading_substring in line
            continue
        if in_section and line.strip().startswith("- "):
            bullets.append(line.strip()[2:].strip())
    return bullets


def compute_validation_summary(
    activity: pd.DataFrame,
    model_status_path: Path = DEFAULT_MODEL_STATUS_PATH,
    validation_results_path: Path = DEFAULT_VALIDATION_RESULTS_PATH,
) -> tuple[pd.DataFrame, list[str]]:
    """Assemble the "Validation Summary" sheet's metrics table and
    limitations list.

    Falls back to "not available" placeholders when the doc files are
    missing (e.g. a test's tmp_path fixture, or a fresh checkout that
    hasn't run `validation/report.py` yet) rather than raising - this
    sheet is a convenience summary, not a gate.
    """
    metrics = compute_validation_metrics(activity)

    validation_results_path = Path(validation_results_path)
    if validation_results_path.exists():
        totals = _extract_validation_totals(validation_results_path.read_text())
    else:
        totals = None
        logger.warning(
            "compute_validation_summary: %s not found, skipping pass/fail totals",
            validation_results_path,
        )
    if totals is not None:
        n_pass, n_fail, n_info = totals
    else:
        n_pass = n_fail = n_info = None
    metrics = pd.concat(
        [
            metrics,
            pd.DataFrame(
                [
                    {"metric": "Validation checks passed", "value": n_pass},
                    {"metric": "Validation checks failed", "value": n_fail},
                    {"metric": "Validation checks informational", "value": n_info},
                ]
            ),
        ],
        ignore_index=True,
    )

    model_status_path = Path(model_status_path)
    if model_status_path.exists():
        limitations = _extract_markdown_bullets(
            model_status_path.read_text(), "Remaining statistical differences"
        )
    else:
        limitations = []
        logger.warning(
            "compute_validation_summary: %s not found, skipping documented limitations",
            model_status_path,
        )
    if not limitations:
        limitations = ["No documented limitations found - see docs/model_status.md directly."]

    return metrics, limitations


# --- Excel writing helpers -------------------------------------------------------


def _infer_number_format(series: pd.Series) -> str | None:
    """A reasonable default Excel number format from a column's dtype -
    integers unformatted-but-thousands-separated, floats to 2 decimals,
    everything else (text, category, bool) left as Excel's General format.
    """
    if pd.api.types.is_bool_dtype(series):
        return None
    if pd.api.types.is_integer_dtype(series):
        return "#,##0"
    if pd.api.types.is_float_dtype(series):
        return "#,##0.00"
    return None


def _autosize_columns(ws: Worksheet, df: pd.DataFrame, start_col: int = 1) -> None:
    for i, column in enumerate(df.columns):
        header_len = len(str(column))
        # Sampling every value's rendered width would be expensive for the
        # largest sheets (synthetic_activity has 15k+ rows); a fixed sample
        # is enough to size a column sensibly without scanning every row.
        sample = df[column].dropna().astype(str).head(200)
        content_len = int(sample.map(len).max()) if not sample.empty else 0
        width = max(MIN_COLUMN_WIDTH, min(MAX_COLUMN_WIDTH, max(header_len, content_len) + 2))
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


def write_table(
    ws: Worksheet,
    df: pd.DataFrame,
    table_name: str,
    style_name: str,
    start_row: int = 1,
    start_col: int = 1,
    time_columns: list[str] | None = None,
    percent_columns: list[str] | None = None,
    column_formats: dict[str, str] | None = None,
) -> int:
    """Write `df` as a formatted, filterable Excel Table starting at
    (`start_row`, `start_col`). Returns the row number immediately after
    the table, for callers stacking more than one table on a sheet.

    Handles this workbook's two recurring special cases: HHMM-encoded time
    columns (`time_columns`, converted via `_hhmm_to_time`) and already-
    computed `datetime.time` columns (detected by dtype - `object` columns
    already holding `dt.time`/`None`, produced by the `compute_*` summary
    functions above, are passed through instead of re-converted).
    """
    time_columns = set(time_columns or [])
    percent_columns = set(percent_columns or [])
    column_formats = column_formats or {}

    n_rows, n_cols = df.shape
    for j, column in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=str(column))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i in range(n_rows):
        for j, column in enumerate(df.columns):
            value = df.iloc[i, j]
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j)

            if column in time_columns:
                cell.value = _hhmm_to_time(value)
                cell.number_format = "h:mm"
                continue
            if isinstance(value, dt.time):
                cell.value = value if pd.notna(value) else None
                cell.number_format = "h:mm"
                continue
            if pd.isna(value):
                cell.value = None
                continue
            if isinstance(value, bool):
                cell.value = value
            elif hasattr(value, "item"):
                cell.value = value.item()
            else:
                cell.value = value

            if column in percent_columns:
                cell.number_format = "0.0%"
            elif column in column_formats:
                cell.number_format = column_formats[column]

    for j, column in enumerate(df.columns):
        if column in time_columns or column in percent_columns or column in column_formats:
            continue
        fmt = _infer_number_format(df[column])
        if fmt is None:
            continue
        for i in range(n_rows):
            ws.cell(row=start_row + 1 + i, column=start_col + j).number_format = fmt

    end_row = start_row + n_rows
    end_col = start_col + n_cols - 1
    if n_rows > 0:
        table_ref = (
            f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        )
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name=style_name, showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)

    _autosize_columns(ws, df, start_col=start_col)
    return end_row + 1


def _write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    table_name: str,
    style_name: str,
    time_columns: list[str] | None = None,
    percent_columns: list[str] | None = None,
    column_formats: dict[str, str] | None = None,
) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    write_table(
        ws,
        df,
        table_name=table_name,
        style_name=style_name,
        time_columns=time_columns,
        percent_columns=percent_columns,
        column_formats=column_formats,
    )
    ws.freeze_panes = "A2"
    return ws


def _write_readme_sheet(
    wb: Workbook,
    style_name: str,
    n_employees: int,
    n_legs: int,
    generated_at: dt.datetime,
) -> Worksheet:
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    title = ws.cell(row=1, column=1, value="Synthetic Mobility Report")
    title.font = Font(bold=True, size=14)

    facts = [
        (
            "Report purpose",
            "Synthetic employee commuting and daily driving activity profiles, "
            "generated for exploratory review by researchers and company "
            "stakeholders without needing Python or Parquet tooling.",
        ),
        ("Generated", generated_at.strftime("%Y-%m-%d %H:%M")),
        ("Number of synthetic employees", n_employees),
        ("Number of activity legs", n_legs),
        ("Source dataset", SOURCE_DATASET_LABEL),
        (
            "Synthetic data",
            "Every row in this workbook is synthetic: statistically resampled "
            "and rescaled from real NHTS respondent travel patterns. It does "
            "not describe any real individual or any real trip.",
        ),
        (
            "EV charging demand",
            "This workbook does not yet include EV ownership assumptions or "
            "workplace charging demand estimates - that is a planned future "
            "modeling stage, not part of this export.",
        ),
    ]
    row = 3
    for label, value in facts:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(vertical="top")
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    sheet_table = pd.DataFrame(SHEET_DESCRIPTIONS, columns=["sheet", "description"])
    write_table(ws, sheet_table, table_name="SheetDirectory", style_name=style_name, start_row=row)

    ws.freeze_panes = "A2"
    return ws


def _write_key_value_sheet(
    wb: Workbook,
    sheet_name: str,
    tables: list[tuple[str, pd.DataFrame, dict]],
    style_names: list[str],
) -> Worksheet:
    """A sheet made of one or more stacked labeled Tables (each preceded by
    a bold section-title row) - used for "Employee Summary" and "Validation
    Summary", which mix a metric/value table with a second differently-
    shaped table (cluster breakdown / limitations list).
    """
    ws = wb.create_sheet(sheet_name)
    row = 1
    for (title, df, kwargs), style_name in zip(tables, style_names):
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        row += 1
        row = write_table(ws, df, style_name=style_name, start_row=row, **kwargs)
        row += 1
    ws.freeze_panes = "A2"
    return ws


# --- Top-level workbook assembly -----------------------------------------------


def build_workbook(
    processed_dir: Path = DEFAULT_PROCESSED_DIR,
    model_status_path: Path = DEFAULT_MODEL_STATUS_PATH,
    validation_results_path: Path = DEFAULT_VALIDATION_RESULTS_PATH,
    generated_at: dt.datetime | None = None,
) -> Workbook:
    """Build the full workbook in memory (no file I/O beyond reading the
    Parquet/doc inputs) so callers - including tests - can inspect it
    before deciding whether/where to save it.
    """
    employees = load_synthetic_employees(processed_dir)
    activity = load_synthetic_activity(processed_dir)
    generated_at = generated_at or dt.datetime.now()

    wb = Workbook()
    wb.remove(wb.active)

    _write_readme_sheet(
        wb,
        style_name=TABLE_STYLES[0],
        n_employees=len(employees),
        n_legs=len(activity),
        generated_at=generated_at,
    )

    employee_export = employees[EMPLOYEE_EXPORT_COLUMNS]
    _write_data_sheet(
        wb,
        "Synthetic Employees",
        employee_export,
        table_name="SyntheticEmployees",
        style_name=TABLE_STYLES[1],
        time_columns=EMPLOYEE_TIME_COLUMNS,
        column_formats=EMPLOYEE_COLUMN_FORMATS,
    )

    activity_export = activity[ACTIVITY_EXPORT_COLUMNS]
    _write_data_sheet(
        wb,
        "Synthetic Activity",
        activity_export,
        table_name="SyntheticActivity",
        style_name=TABLE_STYLES[2],
        time_columns=ACTIVITY_TIME_COLUMNS,
    )

    employee_metrics, cluster_counts = compute_employee_summary(employees, activity)
    _write_key_value_sheet(
        wb,
        "Employee Summary",
        tables=[
            (
                "Population metrics",
                employee_metrics,
                {"table_name": "EmployeeSummaryMetrics"},
            ),
            (
                "Cluster counts and shares",
                cluster_counts,
                {"table_name": "EmployeeSummaryClusters", "percent_columns": ["employee_share"]},
            ),
        ],
        style_names=[TABLE_STYLES[3], TABLE_STYLES[3]],
    )

    cluster_summary = compute_cluster_summary(employees, activity)
    _write_data_sheet(
        wb,
        "Cluster Summary",
        cluster_summary,
        table_name="ClusterSummary",
        style_name=TABLE_STYLES[4],
        percent_columns=["employee_share"],
    )

    hourly_presence = compute_hourly_workplace_presence(activity, len(employees))
    _write_data_sheet(
        wb,
        "Hourly Workplace Presence",
        hourly_presence,
        table_name="HourlyWorkplacePresence",
        style_name=TABLE_STYLES[5],
        percent_columns=["percent_present"],
    )

    validation_metrics, limitations = compute_validation_summary(
        activity, model_status_path, validation_results_path
    )
    limitations_df = pd.DataFrame({"limitation": limitations})
    _write_key_value_sheet(
        wb,
        "Validation Summary",
        tables=[
            (
                "Core validation metrics",
                validation_metrics,
                {"table_name": "ValidationMetrics", "percent_columns": ["value"]},
            ),
            (
                "Important remaining limitations",
                limitations_df,
                {"table_name": "ValidationLimitations"},
            ),
        ],
        style_names=[TABLE_STYLES[6], TABLE_STYLES[6]],
    )

    return wb


def export_all(
    interim_dir: Path | None = None,
    processed_dir: Path = DEFAULT_PROCESSED_DIR,
    reports_dir: Path = DEFAULT_REPORTS_DIR,
    force: bool = False,
    model_status_path: Path = DEFAULT_MODEL_STATUS_PATH,
    validation_results_path: Path = DEFAULT_VALIDATION_RESULTS_PATH,
) -> Path:
    """Pipeline entry point: build the workbook and write it to
    `reports_dir / REPORT_FILENAME`.

    `interim_dir` is accepted (unused) to match `scripts/run_pipeline.py`'s
    shared stage-calling signature - every other export input already
    lives under `processed_dir` or the docs directory. Skips regeneration
    when the output already exists and `force` is not set, mirroring every
    other pipeline stage's existing-output behavior.
    """
    reports_dir = Path(reports_dir)
    output_path = reports_dir / REPORT_FILENAME
    if output_path.exists() and not force:
        logger.info(
            "export_all: %s already exists - skipping (pass force=True to regenerate).",
            output_path,
        )
        return output_path

    wb = build_workbook(processed_dir, model_status_path, validation_results_path)
    reports_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("export_all: wrote %s", output_path)
    return output_path


if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO)

    parser = argparse.ArgumentParser(description="Export the synthetic mobility Excel report.")
    parser.add_argument("--processed-dir", type=Path, default=DEFAULT_PROCESSED_DIR)
    parser.add_argument("--reports-dir", type=Path, default=DEFAULT_REPORTS_DIR)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    export_all(processed_dir=args.processed_dir, reports_dir=args.reports_dir, force=args.force)
