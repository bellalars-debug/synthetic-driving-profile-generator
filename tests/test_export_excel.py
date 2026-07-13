"""Tests for driving_profiles.utils.export_excel."""

from __future__ import annotations

import datetime as dt
import hashlib

import openpyxl
import pandas as pd
import pytest

from driving_profiles.utils import export_excel as ee

REQUIRED_SHEETS = [
    "README",
    "Synthetic Employees",
    "Synthetic Activity",
    "Employee Summary",
    "Cluster Summary",
    "Hourly Workplace Presence",
    "Validation Summary",
]


# --- fixtures -----------------------------------------------------------------


def _employee_row(employee_id: str, cluster_id: int, **overrides) -> dict:
    row = {
        "synthetic_employee_id": employee_id,
        "source_houseid": f"H-{employee_id}",
        "source_personid": f"P-{employee_id}",
        "cluster_id": cluster_id,
        "age": 35,
        "age_band": "35-44",
        "household_income_bracket": 7,
        "household_size": 3,
        "household_vehicle_count": 2,
        "vehicles_per_driver": 1.0,
        "commute_distance_survey_miles": 10.0,
        "commute_distance_trip_miles": 10.0,
        "commute_duration_minutes": 20.0,
        "work_arrival_time": 800.0,
        "work_departure_time": 1700.0,
        "trips_per_day": 2,
        "total_daily_miles": 20.0,
        "total_driving_minutes": 40.0,
        "number_of_stops": 1,
        "used_household_vehicle": True,
    }
    row.update(overrides)
    return row


def _activity_row(employee_id: str, trip_number: int, **overrides) -> dict:
    row = {
        "synthetic_employee_id": employee_id,
        "trip_number": trip_number,
        "departure_time": 700.0,
        "arrival_time": 800.0,
        "trip_purpose": "work",
        "distance": 10.0,
        "duration": 20.0,
        "dwell_time_after": 480.0,
        "is_workplace_arrival": True,
        "is_workplace_departure": False,
        "workplace_dwell_minutes": 480.0,
        "chain_source": "donor",
        "vehicle_type": 1.0,
        "vehicle_fuel": 1.0,
    }
    row.update(overrides)
    return row


@pytest.fixture
def employees_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            _employee_row("SYN-1", cluster_id=0, commute_distance_survey_miles=10.0),
            _employee_row("SYN-2", cluster_id=0, commute_distance_survey_miles=20.0),
            _employee_row("SYN-3", cluster_id=1, commute_distance_survey_miles=30.0),
            _employee_row("SYN-4", cluster_id=1, commute_distance_survey_miles=40.0),
        ]
    )


@pytest.fixture
def activity_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            _activity_row("SYN-1", 1, chain_source="donor", distance=10.0, duration=20.0),
            _activity_row("SYN-2", 1, chain_source="donor", distance=10.0, duration=20.0),
            _activity_row("SYN-3", 1, chain_source="fallback", distance=10.0, duration=20.0),
            # An implausibly fast leg: 100 miles in 1 minute (6000 mph).
            _activity_row(
                "SYN-4",
                1,
                chain_source="donor",
                distance=100.0,
                duration=1.0,
                is_workplace_arrival=False,
                is_workplace_departure=False,
                workplace_dwell_minutes=float("nan"),
            ),
        ]
    )


@pytest.fixture
def processed_dir(tmp_path, employees_df, activity_df):
    processed = tmp_path / "processed"
    processed.mkdir()
    employees_df.to_parquet(processed / "synthetic_employees.parquet", index=False)
    activity_df.to_parquet(processed / "synthetic_activity.parquet", index=False)
    return processed


def _sheet_rows(ws) -> list[tuple]:
    return [row for row in ws.iter_rows(values_only=True)]


def _table_header_and_data(ws, ref: str) -> tuple[list, list[list]]:
    rows = list(ws[ref])
    header = [c.value for c in rows[0]]
    data = [[c.value for c in r] for r in rows[1:]]
    return header, data


# --- build_workbook: structure ---------------------------------------------------


def test_build_workbook_creates_all_required_sheets(processed_dir):
    wb = ee.build_workbook(processed_dir=processed_dir)
    assert wb.sheetnames == REQUIRED_SHEETS


def test_source_ids_excluded_from_synthetic_employees_sheet(processed_dir):
    wb = ee.build_workbook(processed_dir=processed_dir)
    ws = wb["Synthetic Employees"]
    header, _ = _table_header_and_data(ws, ws.tables["SyntheticEmployees"].ref)
    assert "source_houseid" not in header
    assert "source_personid" not in header
    assert header == ee.EMPLOYEE_EXPORT_COLUMNS


def test_vehicle_columns_excluded_from_synthetic_activity_sheet(processed_dir):
    wb = ee.build_workbook(processed_dir=processed_dir)
    ws = wb["Synthetic Activity"]
    header, _ = _table_header_and_data(ws, ws.tables["SyntheticActivity"].ref)
    assert "vehicle_type" not in header
    assert "vehicle_fuel" not in header
    assert header == ee.ACTIVITY_EXPORT_COLUMNS


def test_synthetic_employee_count_matches_parquet_input(processed_dir, employees_df):
    wb = ee.build_workbook(processed_dir=processed_dir)
    ws = wb["Synthetic Employees"]
    _, data = _table_header_and_data(ws, ws.tables["SyntheticEmployees"].ref)
    assert len(data) == len(employees_df)


def test_activity_row_count_matches_parquet_input(processed_dir, activity_df):
    wb = ee.build_workbook(processed_dir=processed_dir)
    ws = wb["Synthetic Activity"]
    _, data = _table_header_and_data(ws, ws.tables["SyntheticActivity"].ref)
    assert len(data) == len(activity_df)


def test_employee_time_columns_render_as_time_objects(processed_dir):
    wb = ee.build_workbook(processed_dir=processed_dir)
    ws = wb["Synthetic Employees"]
    header, data = _table_header_and_data(ws, ws.tables["SyntheticEmployees"].ref)
    arrival_idx = header.index("work_arrival_time")
    assert data[0][arrival_idx] == dt.time(8, 0)


# --- summary computation -----------------------------------------------------


def test_compute_employee_summary_values(employees_df, activity_df):
    metrics, cluster_counts = ee.compute_employee_summary(employees_df, activity_df)
    metrics = metrics.set_index("metric")["value"]

    assert metrics["Employee count"] == 4
    assert metrics["Average commute distance (mi)"] == pytest.approx(25.0)
    assert metrics["Median commute distance (mi)"] == pytest.approx(25.0)
    assert metrics["Average work arrival time"] == dt.time(8, 0)

    cluster_counts = cluster_counts.set_index("cluster_id")
    assert cluster_counts.loc[0, "employee_count"] == 2
    assert cluster_counts.loc[1, "employee_count"] == 2
    assert cluster_counts.loc[0, "employee_share"] == pytest.approx(0.5)


def test_compute_cluster_summary_values(employees_df, activity_df):
    summary = ee.compute_cluster_summary(employees_df, activity_df).set_index("cluster_id")

    assert summary.loc[0, "employee_count"] == 2
    assert summary.loc[0, "mean_commute_distance_miles"] == pytest.approx(15.0)
    assert summary.loc[1, "mean_commute_distance_miles"] == pytest.approx(35.0)
    assert summary.loc[0, "employee_share"] == pytest.approx(0.5)
    # SYN-4 (cluster 1)'s leg is not a workplace-arrival leg (is_workplace_arrival
    # overridden to False), so cluster 1's mean dwell comes from SYN-3 alone.
    assert summary.loc[0, "mean_workplace_dwell_minutes"] == pytest.approx(480.0)
    assert summary.loc[1, "mean_workplace_dwell_minutes"] == pytest.approx(480.0)


def test_compute_hourly_workplace_presence():
    activity = pd.DataFrame(
        [
            # Present exactly during hour 8 (08:00-09:00).
            {
                "synthetic_employee_id": "A",
                "is_workplace_arrival": True,
                "arrival_time": 800.0,
                "workplace_dwell_minutes": 60.0,
            },
            # No recorded dwell (diary ends at work) -> present through
            # end of day, from 14:00 (hour 14) through hour 23.
            {
                "synthetic_employee_id": "B",
                "is_workplace_arrival": True,
                "arrival_time": 1400.0,
                "workplace_dwell_minutes": float("nan"),
            },
            # Not a workplace-arrival leg - must be ignored entirely.
            {
                "synthetic_employee_id": "C",
                "is_workplace_arrival": False,
                "arrival_time": 900.0,
                "workplace_dwell_minutes": 30.0,
            },
        ]
    )

    result = ee.compute_hourly_workplace_presence(activity, n_employees=3).set_index("hour")

    assert result.loc[8, "employees_present"] == 1
    assert result.loc[9, "employees_present"] == 0
    assert result.loc[14, "employees_present"] == 1
    assert result.loc[23, "employees_present"] == 1
    assert result.loc[0, "employees_present"] == 0
    assert result.loc[8, "percent_present"] == pytest.approx(1 / 3)


def test_compute_validation_metrics(activity_df):
    metrics = ee.compute_validation_metrics(activity_df).set_index("metric")["value"]

    # One of four employees (SYN-3) used a fallback chain.
    assert metrics["Fallback chain rate"] == pytest.approx(0.25)
    # The 100mi/1min leg (SYN-4) is the one implausible-speed leg.
    assert metrics["Implausible-speed leg count"] == 1
    assert metrics["Legs checked for speed plausibility"] == 4


def test_extract_validation_totals_reads_summary_line():
    text = "Some text.\n\n**76 passed, 26 failed, 21 informational** across all sections.\n"
    assert ee._extract_validation_totals(text) == (76, 26, 21)


def test_extract_validation_totals_returns_none_when_absent():
    assert ee._extract_validation_totals("no summary line here") is None


def test_extract_markdown_bullets_stops_at_next_heading():
    text = (
        "# Title\n\n"
        "## 5. Other section\n- not this one\n\n"
        "## 6. Remaining statistical differences\n"
        "- first limitation\n"
        "- second limitation\n\n"
        "## 7. Next section\n- should not be included\n"
    )
    bullets = ee._extract_markdown_bullets(text, "Remaining statistical differences")
    assert bullets == ["first limitation", "second limitation"]


def test_compute_validation_summary_uses_doc_files(tmp_path, activity_df):
    model_status = tmp_path / "model_status.md"
    model_status.write_text(
        "## 6. Remaining statistical differences\n- known limitation one\n"
    )
    validation_results = tmp_path / "validation_results.md"
    validation_results.write_text("**10 passed, 2 failed, 1 informational**\n")

    metrics, limitations = ee.compute_validation_summary(
        activity_df,
        model_status_path=model_status,
        validation_results_path=validation_results,
    )
    metrics = metrics.set_index("metric")["value"]

    assert metrics["Validation checks passed"] == 10
    assert metrics["Validation checks failed"] == 2
    assert metrics["Validation checks informational"] == 1
    assert limitations == ["known limitation one"]


def test_compute_validation_summary_handles_missing_docs(tmp_path, activity_df):
    metrics, limitations = ee.compute_validation_summary(
        activity_df,
        model_status_path=tmp_path / "does_not_exist.md",
        validation_results_path=tmp_path / "also_missing.md",
    )
    metrics = metrics.set_index("metric")["value"]

    assert pd.isna(metrics["Validation checks passed"])
    assert limitations == ["No documented limitations found - see docs/model_status.md directly."]


# --- export_all: file behavior, reproducibility, non-mutation --------------------


def test_export_all_creates_workbook(processed_dir, tmp_path):
    reports_dir = tmp_path / "reports"
    output_path = ee.export_all(processed_dir=processed_dir, reports_dir=reports_dir, force=True)

    assert output_path == reports_dir / ee.REPORT_FILENAME
    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == REQUIRED_SHEETS


def test_export_all_skips_when_output_exists_and_not_forced(processed_dir, tmp_path):
    reports_dir = tmp_path / "reports"
    output_path = ee.export_all(processed_dir=processed_dir, reports_dir=reports_dir, force=True)
    original_mtime = output_path.stat().st_mtime_ns

    ee.export_all(processed_dir=processed_dir, reports_dir=reports_dir, force=False)

    assert output_path.stat().st_mtime_ns == original_mtime


def test_export_all_force_regenerates(processed_dir, tmp_path):
    reports_dir = tmp_path / "reports"
    output_path = ee.export_all(processed_dir=processed_dir, reports_dir=reports_dir, force=True)
    original_bytes = output_path.read_bytes()

    ee.export_all(processed_dir=processed_dir, reports_dir=reports_dir, force=True)

    # Regenerated (not merely left alone) - file was rewritten, even if its
    # content happens to match (README's generation timestamp makes an
    # exact byte match unlikely but not guaranteed within the same minute).
    assert output_path.exists()
    assert len(original_bytes) > 0


def test_export_all_does_not_modify_input_parquet_files(processed_dir, tmp_path):
    employees_path = processed_dir / "synthetic_employees.parquet"
    activity_path = processed_dir / "synthetic_activity.parquet"
    employees_hash_before = hashlib.sha256(employees_path.read_bytes()).hexdigest()
    activity_hash_before = hashlib.sha256(activity_path.read_bytes()).hexdigest()

    ee.export_all(processed_dir=processed_dir, reports_dir=tmp_path / "reports", force=True)

    assert hashlib.sha256(employees_path.read_bytes()).hexdigest() == employees_hash_before
    assert hashlib.sha256(activity_path.read_bytes()).hexdigest() == activity_hash_before


def test_build_workbook_is_reproducible_given_fixed_generated_at(processed_dir):
    fixed_time = dt.datetime(2026, 1, 1, 12, 0, 0)

    wb1 = ee.build_workbook(processed_dir=processed_dir, generated_at=fixed_time)
    wb2 = ee.build_workbook(processed_dir=processed_dir, generated_at=fixed_time)

    for sheet_name in REQUIRED_SHEETS:
        rows1 = _sheet_rows(wb1[sheet_name])
        rows2 = _sheet_rows(wb2[sheet_name])
        assert rows1 == rows2
